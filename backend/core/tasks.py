from celery import shared_task
from openpyxl import load_workbook
from .models import Customer, Loan
from datetime import datetime

@shared_task
def load_initial_data():
    # Load customer data
    try:
        wb_cust = load_workbook(filename='customer_data.xlsx')
        ws_cust = wb_cust.active
        for row in ws_cust.iter_rows(min_row=2, values_only=True):
            if not all([row[0], row[1], row[2], row[3], row[4], row[5]]):
                continue  # skip rows missing required data

            Customer.objects.get_or_create(
                customer_id=row[0],
                defaults={
                    'first_name': row[1],
                    'last_name': row[2],
                    'phone_number': str(row[3]),
                    'monthly_salary': int(row[4]),
                    'approved_limit': int(row[5]),
                    'current_debt': float(row[6]) if row[6] is not None else 0.0,
                }
            )
    except Exception as e:
        print(f"Error loading customer data: {e}")

    # Load loan data
    try:
        wb_loan = load_workbook(filename='loan_data.xlsx')
        ws_loan = wb_loan.active
        for row in ws_loan.iter_rows(min_row=2, values_only=True):
            try:
                customer = Customer.objects.get(customer_id=row[0])
                
                # Safely parse start and end dates
                def parse_date(cell):
                    if isinstance(cell, datetime):
                        return cell.date()
                    try:
                        return datetime.strptime(str(cell), "%Y-%m-%d %H:%M:%S").date()
                    except ValueError:
                        return datetime.strptime(str(cell), "%Y-%m-%d").date()

                Loan.objects.get_or_create(
                    loan_id=row[1],
                    defaults={
                        'customer': customer,
                        'loan_amount': float(row[2]),
                        'tenure': int(row[3]),
                        'interest_rate': float(row[4]),
                        'monthly_installment': float(row[5]),
                        'emis_paid_on_time': int(row[6]),
                        'start_date': parse_date(row[7]),
                        'end_date': parse_date(row[8]),
                    }
                )
            except Customer.DoesNotExist:
                print(f"Customer with ID {row[0]} not found. Skipping loan {row[1]}.")
            except Exception as e:
                print(f"Error creating loan for customer {row[0]}: {e}")
    except Exception as e:
        print(f"Error loading loan data: {e}")
